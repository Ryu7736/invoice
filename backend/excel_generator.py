from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime, date
import os
from backend.db import get_invoice_data  # ← 修正: backend.db に変更

def create_invoice_template():
    """請求書のExcelテンプレートを作成"""
    wb = Workbook()
    ws = wb.active
    ws.title = "請求書"
    
    # 🎨 スタイル定義
    title_font = Font(name='MS Gothic', size=16, bold=True)
    header_font = Font(name='MS Gothic', size=12, bold=True)
    normal_font = Font(name='MS Gothic', size=10)
    
    # 📋 請求書ヘッダー
    ws['A1'] = "請求書"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:G1')
    
    # 🏢 会社情報エリア
    ws['A3'] = "発行者:"
    ws['B3'] = "○○タイヤサービス株式会社"
    ws['A4'] = "住所:"
    ws['B4'] = "〒000-0000 東京都○○区○○1-1-1"
    ws['A5'] = "TEL:"
    ws['B5'] = "03-0000-0000"
    
    # 👤 顧客情報エリア
    ws['E3'] = "請求先:"
    ws['F3'] = "{customer_name}"
    ws['E4'] = "住所:"
    ws['F4'] = "{customer_address}"
    ws['E5'] = "TEL:"
    ws['F5'] = "{customer_phone}"
    
    # 📅 請求情報
    ws['A7'] = "請求日:"
    ws['B7'] = "{invoice_date}"
    ws['A8'] = "請求期間:"
    ws['B8'] = "{billing_period}"
    
    # 📊 明細ヘッダー
    headers = ["No.", "商品名", "サイズ/内容", "数量", "単価", "金額", "備考"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # 💰 合計エリア（下部に予約）
    ws['E25'] = "小計:"
    ws['F25'] = "{subtotal}"
    ws['E26'] = "消費税(10%):"
    ws['F26'] = "{tax}"
    ws['E27'] = "合計:"
    ws['F27'] = "{total}"
    
    # セル幅調整
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    return wb

def generate_invoice(customer_id, start_date, end_date, output_dir="./invoices"):
    """指定顧客・期間の請求書を生成"""
    
    # 📁 出力ディレクトリ作成
    os.makedirs(output_dir, exist_ok=True)
    
    # 📊 データ取得
    invoice_data = get_invoice_data(customer_id, start_date, end_date)
    
    if not invoice_data:
        return None, "該当期間のデータがありません"
    
    # 🏭 テンプレート作成
    wb = create_invoice_template()
    ws = wb.active
    
    # 👤 顧客情報を設定
    customer_name = invoice_data[0][0]
    customer_address = invoice_data[0][1] or ""
    customer_phone = invoice_data[0][2] or ""
    
    # 🔄 プレースホルダーを実際の値に置換
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                cell.value = cell.value.replace("{customer_name}", customer_name)
                cell.value = cell.value.replace("{customer_address}", customer_address)
                cell.value = cell.value.replace("{customer_phone}", customer_phone)
                cell.value = cell.value.replace("{invoice_date}", datetime.now().strftime("%Y年%m月%d日"))
                cell.value = cell.value.replace("{billing_period}", f"{start_date} ～ {end_date}")
    
    # 📋 明細データを追加
    subtotal = 0
    row_num = 11  # 明細開始行
    
    for i, data in enumerate(invoice_data, 1):
        product_name = data[3]
        quantity = data[4]
        unit_price = data[5]
        total_amount = data[6]
        tire_size = data[7] or ""
        work_content = data[8] or ""
        
        # 明細行を追加
        ws.cell(row=row_num, column=1, value=i)
        ws.cell(row=row_num, column=2, value=product_name)
        ws.cell(row=row_num, column=3, value=tire_size if tire_size else work_content)
        ws.cell(row=row_num, column=4, value=quantity)
        ws.cell(row=row_num, column=5, value=f"¥{unit_price:,.0f}")
        ws.cell(row=row_num, column=6, value=f"¥{total_amount:,.0f}")
        ws.cell(row=row_num, column=7, value="")
        
        subtotal += total_amount
        row_num += 1
    
    # 💰 合計計算
    tax = int(subtotal * 0.1)
    total = subtotal + tax
    
    # 🔄 合計金額を設定
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                cell.value = cell.value.replace("{subtotal}", f"¥{subtotal:,.0f}")
                cell.value = cell.value.replace("{tax}", f"¥{tax:,.0f}")
                cell.value = cell.value.replace("{total}", f"¥{total:,.0f}")
    
    # 💾 ファイル保存
    filename = f"請求書_{customer_name}_{start_date}_{end_date}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    wb.save(filepath)
    
    return filepath, f"請求書を生成しました: {filename}"

if __name__ == "__main__":
    # テスト実行
    from backend.db import init_database, insert_sample_data  # ← 修正
    
    init_database()
    insert_sample_data()
    
    # テスト用請求書生成
    filepath, message = generate_invoice(1, "2024-01-01", "2024-01-31")
    print(message)
    if filepath:
        print(f"ファイルパス: {filepath}")